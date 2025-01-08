from functools import partial
from http.client import responses
import time
from urllib import response
from xmlrpc.client import ResponseError
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination, LimitOffsetPagination
from . serializers import *
from rest_framework.decorators import api_view,permission_classes, throttle_classes, parser_classes
from rest_framework import status
from rest_framework.permissions import AllowAny
from . models import *
from rest_framework.throttling import UserRateThrottle, AnonRateThrottle , ScopedRateThrottle
from django.conf import settings 
from django.middleware import csrf
from rest_framework.parsers import MultiPartParser
import openpyxl
from rest_framework_simplejwt.views import TokenRefreshView
from django.db import connection
from rest_framework_simplejwt.exceptions import InvalidToken
from . filters import *


def home(request):
    return HttpResponse("Hi There! this is Home :)")


def paginate_my_way(queryset,request,serializer):
    paginator = PageNumberPagination()
    paginator.page_size = 10
    paginated_queryset = paginator.paginate_queryset(queryset.order_by('pk'),request)
    paginator.page_size_query_param = 'size'
    serializer = serializer(paginated_queryset,many=True)
    return paginator.get_paginated_response(serializer.data)


@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([AnonRateThrottle])
def register(request):
    if request.method =="POST":
        serializer = RegistrationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_200_OK)

        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)
    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = CustomLoginLogicSerializer(data=request.data)
    if serializer.is_valid():
        if serializer.context['authorization'] == 'Authorized':
            user = serializer.context['user']
            tokens = get_tokens_for_user(user)
            response = Response({"data":serializer.validated_data,"access_token":tokens['access']},status=status.HTTP_200_OK)

            response.set_cookie(
                key=settings.SIMPLE_JWT['AUTH_COOKIE_REFRESH'],
                value=tokens["refresh"],
                max_age=settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'].total_seconds(),
                secure=settings.SIMPLE_JWT['AUTH_COOKIE_SECURE'],
                httponly=settings.SIMPLE_JWT['AUTH_COOKIE_HTTP_ONLY'],
                samesite=settings.SIMPLE_JWT['AUTH_COOKIE_SAMESITE'],
            )

            # setting the csrf token cookie
            csrf_token = csrf.get_token(request)
            response.set_cookie('csrftoken', csrf_token, samesite='Lax')

            return response
        
        else:
            return Response(status = status.HTTP_401_UNAUTHORIZED)
    
    else:
        return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['POST'])
@permission_classes([AllowAny])
def logout(request):
    refresh_token = request.COOKIES.get('refresh_token')

    if refresh_token:
        try:
            refresh = RefreshToken(refresh_token)
            refresh.blacklist()
        except Exception as e:
            return Response({"error":"There was an error invalidating the token:" + str(e)}, status = status.HTTP_400_BAD_REQUEST)
        
    response = Response(status = status.HTTP_200_OK)
    response.delete_cookie("refresh_token")
    return response


class CookieTokenRefreshView(TokenRefreshView):
    def post(self,request):
        refresh_token = request.COOKIES.get("refresh_token")
        if not refresh_token:
            return Response({"errors":"Refresh token not provided"}, status = status.HTTP_401_UNAUTHORIZED)

        try:
            refresh = RefreshToken(refresh_token)
            access_token = str(refresh.access_token)
            data = {
                "access_token":access_token,
                "expiration_time":settings.SIMPLE_JWT["ACCESS_TOKEN_LIFETIME"].total_seconds()
            }

            return Response({"token_data":data},status = status.HTTP_200_OK)

        except InvalidToken:
            return Response({"errors":"Invalid Token"},status = status.HTTP_401_UNAUTHORIZED)


@api_view(['GET','PUT','PATCH']) #PATCH is for testing purposes
@throttle_classes([UserRateThrottle])
def ProfileManager(request,pk):
    try:
        user_profile = UserProfile.objects.get(user=pk)
    except UserProfile.DoesNotExist:
        return Response({"detail":"Profile does not exist"},status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = UserProfileSerializer(user_profile)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH' # will validate to True if it is 
    serializer = UserProfileSerializer(user_profile,data=request.data,partial=partial)
    if serializer.is_valid():
        serializer.save()
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    else:
        return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET','POST'])
@throttle_classes([UserRateThrottle])
def EventList(request):
    if request.method =="GET":
        filterset = EventFilter(request.GET,queryset=Events.objects.all())
        if not filterset.is_valid():
            return Response({"errors":filterset.errors},status = status.HTTP_400_BAD_REQUEST)
        events = filterset.qs
        count = events.count()
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(events,request,EventSerializer)

    if request.method =='POST':
        serializer = EventSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET','PUT','PATCH','DELETE'])
@throttle_classes([UserRateThrottle])
def EventDetail(request,pk):
    try:
        event = Events.objects.get(id=pk)
    except Events.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = EventSerializer(event)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method =='DELETE':
        event.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH' # will validate to True if it is 
        serializer = EventSerializer(event,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        


@api_view(['GET','POST'])
@throttle_classes([UserRateThrottle])
def AnnouncementList(request):

    if request.user.is_authenticated:
        user = {"role":request.user.role}
    else:
        user = {"role":"Students"}
    
    if request.method =='GET':
        if user.get('role') == "Admin":
            filterset = AnnouncementFilter(request.GET,queryset=Announcements.objects.all())
        else:
            filterset = AnnouncementFilter(request.GET,queryset=Announcements.objects.filter(audiences=user.get('role')))
        if not filterset.is_valid():
            return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)
    
        announcements = filterset.qs
        count = announcements.count()
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(announcements,request,AnnouncementsSerializer)
    
    if request.method == 'POST':
        if user.role != "Admin":
            return Response(status = status.HTTP_401_UNAUTHORIZED)

        serializer = AnnouncementsSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET','PUT','PATCH','DELETE'])
@throttle_classes([UserRateThrottle])
def AnnouncementDetail(request,pk):
    try:
        announcement = Announcements.objects.get(id=pk)
    except Announcements.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = AnnouncementsSerializer(announcement)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH' # will validate to True if it is 
        serializer = AnnouncementsSerializer(announcement,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
    if request.method =='DELETE':
        announcement.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([AllowAny])
def profiles(request):
    user_profiles = UserProfile.objects.all()
    count = user_profiles.count()
    if count == 0:
        return Response({"No profiles found"},status=status.HTTP_204_NO_CONTENT)
    serializer = UserProfileSerializer(user_profiles,many=True)
    return Response({"data":serializer.data},status=status.HTTP_200_OK)
    

@api_view(['GET','PUT','PATCH']) #PATCH is for testing purposes 
@permission_classes([AllowAny])
def SchoolProfileManager(request,pk):
    try:
        school_profile = SchoolProfile.objects.get(user=pk)
    except SchoolProfile.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = SchoolProfileSerializer(school_profile)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH' # will validate to True if it is 
    serializer = SchoolProfileSerializer(school_profile,data=request.data,partial=partial)
    if serializer.is_valid():
        serializer.save()
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    else:
        return Response({"error":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET','POST'])
def SubClassesList(request):
    if request.method =='GET':
        my_subclasses = SubClasses.objects.all()
        serializer = subclassesSerializer(my_subclasses,many=True)
        if my_subclasses.count() == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(my_subclasses,request,subclassesSerializer)
    
    if request.method == 'POST':
        serializer = subclassesSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET','PUT','PATCH','DELETE'])
def SubclassesDetail(request,pk):
    try:
        subclass = SubClasses.objects.get(id=pk)
    except SubClasses.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method=='GET':
        serializer = subclassesSerializer(subclass)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH' 
        serializer = subclassesSerializer(subclass,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        else:
            return Response({"error":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
    if request.method == 'DELETE':
        subclass.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
        

@api_view(['GET'])
def StudentList(request):
    paginator = LimitOffsetPagination()
    paginator.max_limit = 15 
    paginator.default_limit = 3
    if request.method =='GET':
        filterset = StudentFilter(request.GET,queryset=Student.objects.all())
        if not filterset.is_valid():
            return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)
        
        students = filterset.qs
        count = students.count()
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            queryset = paginator.paginate_queryset(students,request)
            serializer = StudentSerializer(queryset,many=True)
            return paginator.get_paginated_response(serializer.data)

    

@api_view(['GET','PUT','PATCH','DELETE'])
def StudentDetail(request,pk):
    try:
        student = Student.objects.get(user=pk)
    except Student.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = StudentSerializer(student)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method =='DELETE':
        student_name = student.firstname
        try:
            student_name_req = request.data['name']
        except Exception as e:
            return Response({"error":f"{e} key must have a value"},status=status.HTTP_400_BAD_REQUEST)
        if student_name == student_name_req:
            student.user.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return Response(status = status.HTTP_401_UNAUTHORIZED)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH'
        serializer = StudentSerializer(student,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['POST'])
@parser_classes([MultiPartParser])
def ImportExport(request):
    if 'file' not in request.FILES:
        return Response({"error": "File is required."}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active  # Assuming data is in the first sheet

        errors = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
            data = {
                "first_name": row[0],
                "last_name": row[1]
            }

            # Validate and save data using the serializer
            serializer = StudentSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
            else:
                errors.append({"row": row_num, "errors": serializer.errors})

        # Check for errors
        if errors:
            return Response({"message": "Some rows failed to import.", "errors": errors},status=status.HTTP_400_BAD_REQUEST)

        return Response({"message": "Users imported successfully."}, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['POST'])
@parser_classes([MultiPartParser])
def ImportStudents(request):
    if 'file' not in request.FILES:
        return Response({"error": "File is required."}, status=status.HTTP_400_BAD_REQUEST)

    file = request.FILES['file']
    try:
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active  # Assuming data is in the first sheet

        data_list = []
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Skip header row
            data_list.append({"first_name": row[0], "last_name": row[1]})

        # Validate and save data using the serializer
        serializer = AddStudentSerializer(data=data_list, many=True)
        if serializer.is_valid():
            # Bulk create the users
            users = serializer.save()

            # Create profiles for each user
            for user in users:
                UserProfile.objects.create(user=user)
                Student.objects.create(user=user)

            return Response({"message": "Students imported successfully."}, status=status.HTTP_201_CREATED)
        else:
            return Response({"message": f"Some rows failed validation. {row_num}", "errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    

@api_view(['GET','POST'])
def ClassesList(request):
    filterset = ClassesFilter(request.GET,queryset=StudentClasses.objects.all())
    if not filterset.is_valid():
        return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)

    if request.method =='GET':
        classes = filterset.qs
        count = classes.count()
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(classes,request,StudentClassSerializer)
        
    if request.method =='POST':
        serializer = StudentClassSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET','PUT','DELETE'])
def ClassesDetail(request,pk):
    try:
        classes = StudentClasses.objects.get(id=pk)
    except StudentClasses.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = StudentClassSerializer(classes)
        return Response(serializer.data,status=status.HTTP_200_OK)
    
    if request.method == 'PUT':
        serializer = StudentClassSerializer(classes,data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data,status=status.HTTP_200_OK)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
    if request.method =='DELETE':
        classes.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    

@api_view(['GET','POST'])
def AcademicList(request):
    filterset = AcademicFilter(request.GET,queryset=Academic.objects.all())
    if not filterset.is_valid():
        return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)
    
    academics = filterset.qs
    if request.method == 'GET':
        count = academics.count()
        if count != 0:
            return paginate_my_way(academics,request,AcademicSerializer)
        else:
            return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method == 'POST':
        serializer = AcademicSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET','PUT','PATCH','DELETE'])
def AcademicDetail(request,pk):
    try:
        academic = Academic.objects.get(id=pk)
    except Academic.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method =='GET':
        serializer = AcademicSerializer(academic)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method =='DELETE':
        academic.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH'
        serializer = AcademicSerializer(academic,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET','POST'])
def SubjectsList(request):
    filterset = SubjectFilter(request.GET,queryset=Subjects.objects.all())
    subjects = filterset.qs
    count = subjects.count()
    
    if request.method == 'GET':
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(subjects,request,SubjectSerializer)
    
    if request.method == 'POST':
        serializer = SubjectSerializer(data = request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET','PUT','PATCH','DELETE'])
def SubjectDetail(request,pk):
    try:
        subject = Subjects.objects.get(id=pk)
    except Subjects.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = SubjectSerializer(subject)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method == 'DELETE':
        subject.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH'
        serializer = SubjectSerializer(subject,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET'])
def StaffList(request):
    paginator = LimitOffsetPagination()
    paginator.max_limit = 15 
    paginator.default_limit = 3
    filterset = StaffFilter(request.GET,queryset=Staff.objects.all())
    if not filterset.is_valid():
        return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)
    
    staff = filterset.qs
    count = staff.count()
    if count == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    elif request.method =='GET':
        queryset = paginator.paginate_queryset(staff,request)
        serializer = StaffSerializer(queryset,many=True)
        return paginator.get_paginated_response(serializer.data)
        

@api_view(['GET','DELETE'])
def StaffDetail(request,pk):
    try:
        staff = Staff.objects.get(user=pk)
    except Staff.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = StaffSerializer(staff)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method == 'DELETE':
        try:
            firstname = request.data['name']
        except Exception as e:
            return Response({"detail":f"{e}"},status=status.HTTP_400_BAD_REQUEST)
        
        if firstname == staff.firstname:
            staff.user.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
    

@api_view(['POST'])
def UserList(request):
    serializer = RoleBasedSerializer(data=request.data,context={"request":request})
    if serializer.is_valid():
        serializer.save()
        return Response(status = status.HTTP_201_CREATED)
    else:
        return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    

@api_view(['GET','PUT','PATCH','DELETE'])
def UserDetail(request,pk):
    try:
        user = CustomUser.objects.get(id=pk)
    except CustomUser.DoesNotExist:
        return Response( status = status.HTTP_404_NOT_FOUND)
    

    if request.method == 'GET':
        serializer = UserSerializer(user)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)

    if request.method == 'DELETE':
        fullname = user.get_full_name()
        try:
            data = request.data['name']
        except Exception as e:
            return Response({"detail":f"{e}"},status=status.HTTP_400_BAD_REQUEST)
        
        if fullname == data:
            user.delete()
            return Response(status = status.HTTP_204_NO_CONTENT)

    if user.approved and user.verified:            
        if request.method in ['PATCH','PUT']:
            partial = request.method == ['PATCH']
            serializer = UserSerializer(user,data=request.data,partial=partial)
            if serializer.is_valid():
                serializer.save()
                return Response({"data":serializer.data},status=status.HTTP_200_OK)
            
            else:
                return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
            
    else:
        return Response(status = status.HTTP_403_FORBIDDEN)
        

@api_view(['POST','PUT'])
def add_ward(request,pk):
    try:
        student = Student.objects.get(user=pk)
    except Student.DoesNotExist:
        return Response(status = status.HTTP_404_NOT_FOUND)
    Parent = Parents.objects.get(user=4) 

    if request.method == 'POST':
        # for the purpose of testing switch out request.user.id
        passkey = request.data['passkey']
        if student.my_passkey == passkey:
            Parent.wards.add(student)
            return Response(status = status.HTTP_200_OK)
        else:
            return Response(status = status.HTTP_400_BAD_REQUEST)
        
    if request.method =='PUT':
        Parent.wards.remove(student)
        return Response(status = status.HTTP_200_OK)

        

@api_view(['GET'])
def ParentList(request):
    paginator = LimitOffsetPagination()
    paginator.max_limit = 15 
    paginator.default_limit = 3
    if request.method =='GET':
        filterset = ParentFilter(request.GET,queryset=Parents.objects.all())
        if not filterset.is_valid():
            return Response({"errors":filterset.errors},status=status.HTTP_400_BAD_REQUEST)
        
        parents = filterset.qs
        count = parents.count()
        if count != 0:
            queryset = paginator.paginate_queryset(parents,request)
            serializer = ParentSerializer(queryset,many=True)
            return paginator.get_paginated_response(serializer.data)
        else:
            return Response(status = status.HTTP_204_NO_CONTENT)
        

@api_view(['GET','PUT','PATCH','DELETE'])
def ParentDetail(request,pk):
    try:
        parent = Parents.objects.get(user=pk)
    except Parents.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = ParentSerializer(parent)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method == 'DELETE':
        parent.user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH'
        serializer = ParentSerializer(parent,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET','POST'])
def RecordList(request):
    #user = request.user
    user = {"role":"Admin"}
    paginator = LimitOffsetPagination()
    paginator.max_limit = 15 
    paginator.default_limit = 3


    if user.get('role') != "Parent":
        filterset = RecordFilter(request.GET, queryset=Assessment_records.objects.all())
        if not filterset.is_valid():
            return Response({"errors": filterset.errors}, status=status.HTTP_400_BAD_REQUEST)
        records = filterset.qs


    else:
        try:
            parent = Parents.objects.get(user=2)
        except Parents.DoesNotExist:
            return Response(status=status.HTTP_404_NOT_FOUND)
        

        student_ids = parent.wards.values_list('user', flat=True)
        filterset = RecordFilter(request.GET, queryset=Assessment_records.objects.filter(student__in=student_ids))
        if not filterset.is_valid():
            return Response({"errors": filterset.errors}, status=status.HTTP_400_BAD_REQUEST)
        records = filterset.qs


    if request.method == 'GET':
        count = records.count()
        if count == 0:
            return Response(status=status.HTTP_204_NO_CONTENT)
        
        queryset = paginator.paginate_queryset(records,request)
        serializer = RecordSerializer(queryset,many=True)
        return paginator.get_paginated_response(serializer.data)


    if request.method == 'POST':
        serializer = RecordSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response({"errors": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
    
@api_view(['GET','PUT','PATCH','DELETE'])
def RecordDetail(request,pk):
    try:
        record = Assessment_records.objects.get(id=pk)
    except Assessment_records.DoesNotExist:
        return Response(status = status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = RecordSerializer(record)
        return Response({"data":serializer.data},status=status.HTTP_200_OK)
    
    if request.method == 'DELETE':
        record.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PATCH','PUT']:
        partial = request.method == 'PATCH'
        serializer = RecordSerializer(record,data=request.data,partial=partial)
        if serializer.is_valid():
            serializer.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['POST'])
@throttle_classes([ScopedRateThrottle])
def ChangeMail(request):
    ChangeMail.throttle_scope = 'change_mail'
    
    value = request.data.get('id')
    email = request.data.get('email')
    otp = request.data.get('otp')
    if not value or not email:
        return Response(status = status.HTTP_400_BAD_REQUEST)
    
    if not otp:
        return Response(status= status.HTTP_400_BAD_REQUEST)

    try:
        user = CustomUser.objects.get(id=value)
    except CustomUser.DoesNotExist:
        return Response(status = status.HTTP_404_NOT_FOUND)
    
    if CustomUser.objects.filter(email=email).exists():
        return Response(status=status.HTTP_409_CONFLICT)

    old_email = user.email
    user.email = email
    user.save()
    mail = MailChange()
    mail.user = user
    mail.is_verified = True
    mail.verification_token = otp
    mail.old_email = old_email
    mail.new_email = email
    mail.save()
    return Response(status=status.HTTP_200_OK)



@api_view(['GET'])
def ApprovalsList(request):
    if request.method == 'GET':
        filterset = ApprovalsFilter(request.GET,queryset=Approvals.objects.all())
        approvals = filterset.qs
        count = approvals.count()
        if count == 0:
            return Response(status = status.HTTP_204_NO_CONTENT)
        else:
            return paginate_my_way(approvals,request,ApprovalsSerializer)
        

@api_view(['GET','PUT','PATCH','DELETE'])
def ApprovalsDetail(request,pk):
    try:
        approval = Approvals.objects.get(id=pk)
    except Approvals.DoesNotExist:
        return Response(status = status.HTTP_404_NOT_FOUND)

    if request.method == 'GET':
        serializer = ApprovalsSerializer(approval)
        return Response({"data":serializer.data},status=status.HTTP_200_OK) 
    
    if request.method == 'DELETE':
        approval.user.delete()
        return Response(status = status.HTTP_204_NO_CONTENT)
    
    if request.method in ['PUT']:
        serializer = ApprovalsSerializer(approval,data=request.data)
        if serializer.is_valid():
            serializer.save()
            user = CustomUser.objects.get(id=approval.user)
            user.approved = True
            user.save()
            return Response({"data":serializer.data},status=status.HTTP_200_OK)
        else:
            return Response({"errors":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
        

@api_view(['GET'])
def SchoolDetail(request):
    schema_name = connection.schema_name # this gets the schema name
    subdomain_format = F"https://www.{schema_name}.com"
    try:
        school_profile = SchoolProfile.objects.get(subdomain_url=subdomain_format)
    except SchoolProfile.DoesNotExist:
        return Response(status = status.HTTP_404_NOT_FOUND)
    
    serializer = SchoolProfileSerializer(school_profile)
    return Response(
        {"data":serializer.data},
        status = status.HTTP_200_OK
    )


@api_view(['GET','POST'])
def TransactionsList(request):
    pass


@api_view(['GET'])
def TransactionDetail(request,pk):
    pass 


@api_view(['POST'])
def VerifyMail(request):
    if request.method == 'POST':
        value = request.data.get('id')
        if not value:
            return Response(status = status.HTTP_400_BAD_REQUEST)
        
        try:
            user = CustomUser.objects.get(id=value)
        except CustomUser.DoesNotExist:
            return Response(status = status.HTTP_404_NOT_FOUND)
        
        if user:
            user.verfied = True
            user.save()
            return Response( status = status.HTTP_200_OK)